import os
import pandas as pd
from docx import Document
from PyPDF2 import PdfReader
from openpyxl import load_workbook
from flask import render_template, request, redirect, url_for, session
from app import app, db
from app.models import Question, UserAnswer, WrongSet


@app.route('/')
def index():
    if 'user_id' not in session:
        return redirect(url_for('login'))
    return render_template('index.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        # 简单的登录验证，这里可以扩展为数据库验证
        session['user_id'] = 1
        return redirect(url_for('index'))
    return render_template('login.html')


@app.route('/upload', methods=['POST'])
def upload():
    file = request.files['file']
    if file:
        file_path = os.path.join('app/upload', file.filename)
        file.save(file_path)
        if file.filename.endswith('.docx'):
            process_docx(file_path, file.filename)
        elif file.filename.endswith('.pdf'):
            process_pdf(file_path, file.filename)
        elif file.filename.endswith('.xlsx'):
            process_excel(file_path, file.filename)
    return redirect(url_for('index'))


def process_docx(file_path, file_name):
    doc = Document(file_path)
    question = ""
    question_type = ""
    options = ""
    answer = ""
    image_path = ""
    category = ""
    for para in doc.paragraphs:
        text = para.text
        if text.startswith('题目类型'):
            question_type = text.split(':')[1].strip()
        elif text.startswith('题目'):
            question = text.split(':')[1].strip()
        elif text.startswith('选项'):
            options = text.split(':')[1].strip()
        elif text.startswith('答案'):
            answer = text.split(':')[1].strip()
        elif text.startswith('图片路径'):
            image_path = text.split(':')[1].strip()
        elif text.startswith('类别'):
            category = text.split(':')[1].strip()
        if question and question_type and options and answer:
            new_question = Question(question_text=question, question_type=question_type, options=options, answer=answer,
                                    image_path=image_path, category=category, file_name=file_name)
            db.session.add(new_question)
            db.session.commit()
            question = ""
            question_type = ""
            options = ""
            answer = ""
            image_path = ""
            category = ""


def process_pdf(file_path, file_name):
    reader = PdfReader(file_path)
    text = ""
    for page in reader.pages:
        text += page.extract_text()
    lines = text.split('\n')
    question = ""
    question_type = ""
    options = ""
    answer = ""
    image_path = ""
    category = ""
    for line in lines:
        if line.startswith('题目类型'):
            question_type = line.split(':')[1].strip()
        elif line.startswith('题目'):
            question = line.split(':')[1].strip()
        elif line.startswith('选项'):
            options = line.split(':')[1].strip()
        elif line.startswith('答案'):
            answer = line.split(':')[1].strip()
        elif line.startswith('图片路径'):
            image_path = line.split(':')[1].strip()
        elif line.startswith('类别'):
            category = line.split(':')[1].strip()
        if question and question_type and options and answer:
            new_question = Question(question_text=question, question_type=question_type, options=options, answer=answer,
                                    image_path=image_path, category=category, file_name=file_name)
            db.session.add(new_question)
            db.session.commit()
            question = ""
            question_type = ""
            options = ""
            answer = ""
            image_path = ""
            category = ""


def process_excel(file_path, file_name):
    wb = load_workbook(file_path)
    sheet = wb.active
    for row in sheet.iter_rows(values_only=True):
        question_type, question, options, answer, image_path, category = row
        new_question = Question(question_text=question, question_type=question_type, options=options, answer=answer,
                                image_path=image_path, category=category, file_name=file_name)
        db.session.add(new_question)
        db.session.commit()


@app.route('/start_quiz', methods=['GET'])
def start_quiz():
    category = request.args.get('category')
    order = request.args.get('order', 'asc')
    questions = Question.query.filter_by(category=category).order_by(
        Question.id.asc() if order == 'asc' else Question.id.desc()).all()
    session['questions'] = [{'id': q.id, 'text': q.question_text, 'type': q.question_type, 'options': q.options,
                             'answer': q.answer, 'image_path': q.image_path} for q in questions]
    session['current_question'] = 0
    session['correct_count'] = 0
    session['total_questions'] = len(questions)
    return redirect(url_for('question'))


@app.route('/question')
def question():
    if 'user_id' not in session or 'questions' not in session or 'current_question' not in session:
        return redirect(url_for('index'))
    if session['current_question'] >= len(session['questions']):
        return "答题已结束", 400
    current_question = session['questions'][session['current_question']]
    return render_template('question.html', question=current_question)

@app.route('/answer', methods=['POST'])
def answer():
    user_answer = request.form.get('answer')
    question_id = session['questions'][session['current_question']]['id']
    correct_answer = session['questions'][session['current_question']]['answer']
    is_correct = user_answer == correct_answer
    new_answer = UserAnswer(user_id=session['user_id'], question_id=question_id, user_answer=user_answer,
                            is_correct=is_correct)
    db.session.add(new_answer)
    if not is_correct:
        new_wrong_set = WrongSet(user_id=session['user_id'], question_id=question_id)
        db.session.add(new_wrong_set)
    db.session.commit()
    session['current_question'] += 1
    if session['current_question'] >= len(session['questions']):
        # 处理答题结束的逻辑，例如重定向到结果页面
        session['correct_count'] = sum([answer['is_correct'] for answer in session['questions']])
        return redirect(url_for('result'))
    return redirect(url_for('question'))


@app.route('/result')
def result():
    if 'user_id' not in session or 'correct_count' not in session or 'total_questions' not in session:
        return redirect(url_for('index'))
    correct_count = session['correct_count']
    total_questions = session['total_questions']
    return render_template('result.html', correct_count=correct_count, total_questions=total_questions)


@app.route('/wrong_set')
def wrong_set_view():
    if 'user_id' not in session:
        return redirect(url_for('index'))
    user_id = session['user_id']
    wrong_set_questions = WrongSet.query.filter_by(user_id=user_id).join(Question).all()
    return render_template('wrong_set.html', wrong_set_questions=wrong_set_questions)